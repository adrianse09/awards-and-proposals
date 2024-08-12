#Total Code
import pandas as pd
from openpyxl import Workbook, load_workbook

# Specify the correct file paths
file_path_awards = "C:\\Users\\abuitronboada\\Box\\CombineNode\\Jupyter\\Awards_report.xlsx"
file_path_proposals = "C:\\Users\\abuitronboada\\Box\\CombineNode\\Jupyter\\Proposals_report.xlsx"

# Read the Excel files into DataFrames
awards_df = pd.read_excel(file_path_awards)
proposals_df = pd.read_excel(file_path_proposals)

# Create a Workbook object
workbook = Workbook()

# Create a worksheet for awards
worksheet_awards = workbook.active
worksheet_awards.title = "Awards"

# Create a worksheet for proposals
worksheet_proposals = workbook.create_sheet(title="Proposals")

# Specify the titles for the Awards sheet
awards_titles = [
    "Project Number",
    "Award Lead Pi",
    "Project Title",
    "Sponsor",
    "Award Amount",
    "Award Notice Date"
]

# Specify the titles for the Proposals sheet
proposals_titles = [
    "Proposal Number",
    "Proposal Lead Pi Name",
    "Proposal Specialist",
    "Owner",
    "Proposal Title",
    "Proposal Sponsor",
    "Proposal Type",
    "Proposal Status",
    "Submitted to Sponsor",
    "Total Amount Proposed"
]

# Write column titles to the Awards sheet
worksheet_awards.append(awards_titles)

# Write column titles to the Proposals sheet
worksheet_proposals.append(proposals_titles)

# Write data from awards_df to the Awards sheet
for idx, row in awards_df.iterrows():
    worksheet_awards.append(list(row))

# Write data from proposals_df to the Proposals sheet
for idx, row in proposals_df.iterrows():
    worksheet_proposals.append(list(row))

# Save the workbook
workbook.save(r'C:\Users\adria\Desktop\combined.xlsx')
print("Workbook saved as 'combined.xlsx' on your Desktop.")


# Load the workbook to calculate the total award amount
workbook = load_workbook('C:\Users\abuitronboada\Box\CombineNode\Jupyter\combined.xlsx')
worksheet_awards = workbook['Awards']
worksheet_proposals = workbook['Proposals']

# Initialize variables to store the total amounts
total_award_amount = 0
proposals_total = 0

# Iterate over the rows in the "Awards" sheet starting from the second row
for row in worksheet_awards.iter_rows(min_row=2, values_only=True):
    award_amount = row[4]  # Assuming "Award Amount" is the fifth column (index 4)
    if isinstance(award_amount, (int, float)):
        total_award_amount += award_amount

# Iterate over the rows in the "Proposals" sheet starting from the second row
for row in worksheet_proposals.iter_rows(min_row=2, values_only=True):
    proposals_amount = row[-1]  # Assuming "Total Amount Proposed" is the last column
    if isinstance(proposals_amount, (int, float)):
        proposals_total += proposals_amount

# Print the total amounts
print("Total Award Amount:", total_award_amount)
print("Total Proposals Amount:", proposals_total)

# Load the Awards sheet from the combined.xlsx file
combined_awards_df = pd.read_excel('C:\\Users\\abuitronboada\\Box\\CombineNode\\Jupyter\\combined.xlsx', sheet_name='Awards')

# Load the Proposals sheet from the combined.xlsx file
combined_proposals_df = pd.read_excel('C:\\Users\\abuitronboada\\Box\\CombineNode\\Jupyter\\combined.xlsx', sheet_name='Proposals')

# Display the first few rows of each sheet
print("Awards Sheet:")
display(combined_awards_df.head())

print("\nProposals Sheet:")
display(combined_proposals_df.head())
